from reportlab.pdfgen import canvas
from django.http import FileResponse

def analytics_inventory_pdf(request):
    # Example: generate a simple PDF with product summary
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=inventory_analytics.pdf'
    p = canvas.Canvas(response)
    p.setFont("Helvetica", 14)
    p.drawString(100, 800, "Inventory Analytics Report")
    y = 780
    for prod in Product.objects.all()[:50]:
        p.setFont("Helvetica", 10)
        therapeutic = prod.get_therapeutic_class_display() if hasattr(prod, 'get_therapeutic_class_display') else getattr(prod, 'therapeutic_class', '')
        qty = getattr(prod, 'current_stock', 'N/A')
        p.drawString(100, y, f"{prod.name} | {therapeutic} | Qty: {qty}")
        y -= 16
        if y < 50:
            p.showPage()
            y = 800
    p.save()
    return response
# --- Excel Export/Import ---
import openpyxl
from django.http import HttpResponse
from .models import Product
from django.contrib import messages
from django.shortcuts import redirect

def export_products_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["ID", "Name", "DCI", "Therapeutic Class", "Cost Price", "Selling Price", "Is Active", "Barcode", "Current Stock", "Minimum Stock Level"])
    for p in Product.objects.all():
        # Get human-friendly therapeutic class where possible
        therapeutic = p.get_therapeutic_class_display() if hasattr(p, 'get_therapeutic_class_display') else getattr(p, 'therapeutic_class', '')
        ws.append([
            p.id,
            p.name,
            getattr(p, 'dci', ''),
            therapeutic,
            float(p.cost_price) if p.cost_price is not None else 0,
            float(getattr(p, 'selling_price', getattr(p, 'selling_price', 0))) if getattr(p, 'selling_price', None) is not None else 0,
            bool(p.is_active),
            p.barcode or '',
            int(getattr(p, 'current_stock', 0)),
            int(getattr(p, 'minimum_stock_level', 0)),
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products_export.xlsx'
    wb.save(response)
    return response


def download_products_template(request):
    """Return a small Excel template for product imports."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products Template"
    headers = ["ID", "Name", "DCI", "Therapeutic Class", "Cost Price", "Selling Price", "Is Active", "Barcode", "Current Stock", "Minimum Stock Level"]
    ws.append(headers)
    # example row
    ws.append(["", "Example Product", "DCI Example", "other", 100.0, 150.0, True, "", 10, 5])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products_template.xlsx'
    wb.save(response)
    return response

def import_products_excel(request):
    if request.method != 'POST' or not request.FILES.get('file'):
        messages.error(request, 'Aucun fichier envoyé. Veuillez sélectionner un fichier Excel (.xlsx).')
        return redirect('product_list')

    wb = openpyxl.load_workbook(request.FILES['file'])
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    successful = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        # Ensure we have 10 columns, pad if necessary
        fields = list(row)
        while len(fields) < 10:
            fields.append(None)
        _id, name, dci, therapeutic_class, cost_price, selling_price, is_active, barcode, current_stock, minimum_stock = fields

        # Basic validations
        row_errors = []
        if not name or str(name).strip() == '':
            row_errors.append('Name is required')

        # Parse numeric fields
        try:
            cost_price_val = float(cost_price) if cost_price not in (None, '') else 0.0
        except Exception:
            row_errors.append('Invalid cost price')
            cost_price_val = 0.0

        try:
            selling_price_val = float(selling_price) if selling_price not in (None, '') else 0.0
        except Exception:
            row_errors.append('Invalid selling price')
            selling_price_val = 0.0

        try:
            current_stock_val = int(current_stock) if current_stock not in (None, '') else 0
        except Exception:
            row_errors.append('Invalid current stock')
            current_stock_val = 0

        try:
            minimum_stock_val = int(minimum_stock) if minimum_stock not in (None, '') else 5
        except Exception:
            row_errors.append('Invalid minimum stock level')
            minimum_stock_val = 5

        # Normalize therapeutic class
        therapeutic_val = (therapeutic_class or 'other')
        # Validate therapeutic class against allowed choices
        allowed_classes = [c[0] for c in Product.THERAPEUTIC_CLASSES]
        if therapeutic_val not in allowed_classes:
            row_errors.append(f'Invalid therapeutic class: {therapeutic_val}')

        if row_errors:
            errors.append({'row': idx, 'errors': row_errors})
            continue

        defaults = {
            'name': str(name).strip(),
            'dci': str(dci).strip() if dci else '',
            'therapeutic_class': therapeutic_val,
            'cost_price': cost_price_val,
            'selling_price': selling_price_val,
            'is_active': bool(is_active) if is_active not in (None, '') else True,
            'current_stock': current_stock_val,
            'minimum_stock_level': minimum_stock_val,
        }

        try:
            if barcode and str(barcode).strip() != '':
                Product.objects.update_or_create(barcode=str(barcode).strip(), defaults=defaults)
            else:
                Product.objects.update_or_create(name=defaults['name'], defaults=defaults)
            successful += 1
        except Exception as e:
            errors.append({'row': idx, 'errors': [f'Database error: {e}']})

    # Prepare messages
    if successful:
        messages.success(request, f'{successful} produit(s) importé(s) avec succès.')
    if errors:
        # attach up to first 10 error summaries
        for err in errors[:10]:
            messages.error(request, f"Ligne {err['row']}: {', '.join(err['errors'])}")
        if len(errors) > 10:
            messages.error(request, f"...et {len(errors)-10} autres erreurs.")

    return redirect('product_list')

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import JsonResponse, Http404
from django.utils import timezone
from datetime import timedelta
from .models import Product, ProductBatch
from .forms import ProductForm, ProductBatchForm
# Supplier and PurchaseOrder features removed from views (hidden/deleted)

from django.db.models import Sum, Avg, Count, F, Q, ExpressionWrapper, DecimalField
from datetime import datetime, timedelta
import json
from sales.models import SaleItem, Sale

def analytics_dashboard(request):
    return render(request, 'inventory/analytics_dashboard.html')

def sales_analytics_api(request):
    # Date range (last 30 days by default)
    days = int(request.GET.get('days', 30))
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)

    # Sales data
    sales_data = Sale.objects.filter(
        sale_date__date__gte=start_date,
        sale_date__date__lte=end_date
    )

    # Key metrics
    total_sales = sales_data.count()
    total_revenue = sales_data.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    average_sale = sales_data.aggregate(Avg('total_amount'))['total_amount__avg'] or 0

    # Sales by day for chart
    sales_by_day = sales_data.extra(
        {'sale_day': "date(sale_date)"}
    ).values('sale_day').annotate(
        daily_sales=Count('id'),
        daily_revenue=Sum('total_amount')
    ).order_by('sale_day')

    # Payment type distribution
    payment_distribution = sales_data.values('sale_type').annotate(
        count=Count('id'),
        revenue=Sum('total_amount')
    )

    # Top selling products
    top_products = SaleItem.objects.filter(
        sale__sale_date__date__gte=start_date,
        sale__sale_date__date__lte=end_date
    ).values('product__name').annotate(
        quantity_sold=Sum('quantity'),
        revenue=Sum(F('quantity') * F('unit_price'))
    ).order_by('-quantity_sold')[:10]

    return JsonResponse({
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'days': days
        },
        'metrics': {
            'total_sales': total_sales,
            'total_revenue': float(total_revenue),
            'average_sale': float(average_sale),
            'sales_per_day': round(total_sales / days, 1) if days > 0 else 0,
            'revenue_per_day': float(total_revenue / days) if days > 0 else 0
        },
        'charts': {
            'sales_by_day': list(sales_by_day),
            'payment_distribution': list(payment_distribution),
            'top_products': list(top_products)
        }
    })

def profitability_analytics_api(request):
    # Calculate profitability for products - FIXED VERSION
    products_with_profit = []
    for product in Product.objects.filter(is_active=True, cost_price__gt=0):
        if product.cost_price > 0:
            profit_margin = ((product.selling_price - product.cost_price) / product.cost_price) * 100
            total_profit = (product.selling_price - product.cost_price) * product.current_stock
        else:
            profit_margin = 0
            total_profit = 0

        products_with_profit.append({
            'name': product.name,
            'profit_margin': float(profit_margin),
            'total_profit': float(total_profit),
            'selling_price': float(product.selling_price),
            'cost_price': float(product.cost_price),
            'current_stock': product.current_stock
        })

    # Sort by profit margin (highest first)
    products_with_profit.sort(key=lambda x: x['profit_margin'], reverse=True)

    # Sales-based profitability (last 30 days)
    thirty_days_ago = timezone.now().date() - timedelta(days=30)
    product_profits = SaleItem.objects.filter(
        sale__sale_date__date__gte=thirty_days_ago
    ).annotate(
        product_cost_price=F('product__cost_price')
    ).annotate(
        profit_per_item=ExpressionWrapper(
            F('unit_price') - F('product_cost_price'),
            output_field=DecimalField()
        ),
        total_profit=ExpressionWrapper(
            (F('unit_price') - F('product_cost_price')) * F('quantity'),
            output_field=DecimalField()
        )
    ).values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('unit_price')),
        total_profit=Sum('total_profit')
    ).order_by('-total_profit')[:15]

    # Overall profitability metrics
    total_inventory_value = sum(
        product.cost_price * product.current_stock
        for product in Product.objects.filter(is_active=True)
    )
    total_potential_revenue = sum(
        product.selling_price * product.current_stock
        for product in Product.objects.filter(is_active=True)
    )
    total_potential_profit = total_potential_revenue - total_inventory_value

    return JsonResponse({
        'profitability_metrics': {
            'total_inventory_value': float(total_inventory_value),
            'total_potential_revenue': float(total_potential_revenue),
            'total_potential_profit': float(total_potential_profit),
            'overall_margin_percentage': float(
                (total_potential_profit / total_inventory_value * 100)
                if total_inventory_value > 0 else 0
            )
        },
        'most_profitable_products': products_with_profit[:10],  # Top 10 by profit margin
        'sales_based_profits': list(product_profits)
    })
def inventory_analytics_api(request):
    # Inventory turnover analysis
    products = Product.objects.filter(is_active=True)

    # Stock status breakdown
    stock_status = {
        'out_of_stock': products.filter(current_stock=0).count(),
        'low_stock': products.filter(current_stock__lte=F('minimum_stock_level'), current_stock__gt=0).count(),
        'in_stock': products.filter(current_stock__gt=F('minimum_stock_level')).count()
    }

    # Value analysis
    total_inventory_value = sum(p.cost_price * p.current_stock for p in products)

    # Fast/slow moving analysis (based on recent sales)
    thirty_days_ago = timezone.now().date() - timedelta(days=30)
    product_movement = SaleItem.objects.filter(
        sale__sale_date__date__gte=thirty_days_ago
    ).values('product__name', 'product__current_stock').annotate(
        units_sold=Sum('quantity')
    ).order_by('-units_sold')

    # Expiry analysis - FIXED VERSION
    today = timezone.now().date()
    expiry_alerts = []
    # Only consider batches that actually have quantity and belong to active products
    for batch in ProductBatch.objects.filter(
        expiry_date__gte=today,
        quantity__gt=0,
        product__is_active=True
    ).order_by('expiry_date')[:10]:
        days_until_expiry = (batch.expiry_date - today).days
        expiry_alerts.append({
            'product__name': batch.product.name,
            'expiry_date': batch.expiry_date.strftime('%Y-%m-%d'),
            'days_until_expiry': days_until_expiry,
            'quantity': batch.quantity
        })

    return JsonResponse({
        'stock_status': stock_status,
        'inventory_value': float(total_inventory_value),
        'total_products': products.count(),
        'product_movement': list(product_movement[:15]),
        'expiry_alerts': expiry_alerts,
        'average_stock_level': float(sum(p.current_stock for p in products) / products.count()) if products.count() > 0 else 0
    })
# Product List for staff
def product_list(request):
    products = Product.objects.filter(is_active=True).order_by('name')
    return render(request, 'inventory/product_list.html', {'products': products})

# Add/Edit Product
def product_edit(request, product_id=None):
    if product_id:
        product = get_object_or_404(Product, id=product_id)
    else:
        product = None

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, f'Produit {"modifié" if product else "ajouté"} avec succès!')
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)

    return render(request, 'inventory/product_edit.html', {
        'form': form,
        'product': product
    })

# Receive Stock (Add Batch)
def receive_stock(request, product_id=None):
    product = None
    if product_id:
        product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = ProductBatchForm(request.POST)
        if form.is_valid():
            batch = form.save()
            # Update product stock
            product = batch.product
            product.current_stock += batch.quantity
            product.save()

            messages.success(request, f'Stock reçu: {batch.quantity} unités de {product.name}')
            return redirect('inventory_dashboard')
    else:
        initial = {'product': product} if product else {}
        form = ProductBatchForm(initial=initial)

    return render(request, 'inventory/receive_stock.html', {
        'form': form,
        'product': product
    })

# Stock Adjustment
def adjust_stock(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        adjustment = int(request.POST.get('adjustment', 0))
        reason = request.POST.get('reason', '')

        if adjustment != 0:
            product.current_stock += adjustment
            if product.current_stock < 0:
                product.current_stock = 0
            product.save()

            # Log the adjustment (we'll create a model for this later)
            messages.success(request, f'Stock ajusté: {adjustment} unités. Stock actuel: {product.current_stock}')
            return redirect('product_list')

    return render(request, 'inventory/adjust_stock.html', {'product': product})

def inventory_dashboard(request):
    return render(request, 'inventory/dashboard.html')

def inventory_dashboard_api(request):
    # Get all active products
    products = Product.objects.filter(is_active=True).select_related()

    # Calculate dashboard metrics
    total_products = products.count()
    total_stock_value = sum(product.total_value for product in products)

    # Stock alerts
    low_stock_products = [p for p in products if p.stock_status == 'low_stock']
    out_of_stock_products = [p for p in products if p.stock_status == 'out_of_stock']

    # Expiry alerts
    today = timezone.now().date()
    # Only count batches with remaining quantity and active products
    critical_expiry = ProductBatch.objects.filter(
        expiry_date__lte=today + timedelta(days=30),
        expiry_date__gte=today,
        quantity__gt=0,
        product__is_active=True
    ).count()
    expired_batches = ProductBatch.objects.filter(
        expiry_date__lt=today,
        quantity__gt=0,
        product__is_active=True
    ).count()

    # Get critical expiry batches for display
    critical_batches = ProductBatch.objects.filter(
        expiry_date__lte=today + timedelta(days=30),
        expiry_date__gte=today,
        quantity__gt=0,
        product__is_active=True
    ).select_related('product')[:10]

    # Prepare product data for the table
    product_data = []
    for product in products:
        product_data.append({
            'id': product.id,
            'name': product.name,
            'dci': product.dci,
            'therapeutic_class': product.get_therapeutic_class_display(),
            'current_stock': product.current_stock,
            'minimum_stock_level': product.minimum_stock_level,
            'stock_status': product.stock_status,
            'stock_status_display': product.get_stock_status_display(),
            'selling_price': float(product.selling_price),
            'cost_price': float(product.cost_price),
            'total_value': float(product.total_value),
        })

    # Sort products by stock status (out of stock first, then low stock)
    product_data.sort(key=lambda x: (
        0 if x['stock_status'] == 'out_of_stock' else
        1 if x['stock_status'] == 'low_stock' else 2
    ))

    return JsonResponse({
        'metrics': {
            'total_products': total_products,
            'total_stock_value': float(total_stock_value),
            'low_stock_count': len(low_stock_products),
            'out_of_stock_count': len(out_of_stock_products),
            'critical_expiry_count': critical_expiry,
            'expired_count': expired_batches,
        },
        'alerts': {
            'low_stock': [
                {
                    'name': p.name,
                    'current_stock': p.current_stock,
                    'minimum_stock': p.minimum_stock_level
                }
                for p in low_stock_products
            ],
            'out_of_stock': [
                {
                    'name': p.name,
                    'current_stock': p.current_stock
                }
                for p in out_of_stock_products
            ]
        },
        'products': product_data
    })

# Supplier and PurchaseOrder UI removed — features intentionally deleted from views
